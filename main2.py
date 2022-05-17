from openpyxl import load_workbook
import xml.dom.minidom
import os
import tkinter.messagebox

# 输入输出文件的路径
inputFilePath = '颜色规范.xlsx'
outPutFileDir = "yowa_color_output/"

def mkdir(path):
    folder = os.path.exists(path)
    if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
        os.makedirs(path)  # makedirs 创建文件时如果路径不存在会创建这个路径
        print
        "---  new folder...  ---"
        print
        "---  OK  ---"
    else:
        print
        "---  There is this folder!  ---"

class ColumnIndexStore:
    keyIndex = -1
    dayValueIndex = -1
    nightValueIndex = -1

def findSheetColumnIndex(sheet):
    #取第一行的数据
    cells = list(sheet.values)[0]
    currentIndex = 1
    columnIndexStore = ColumnIndexStore()
    for cell in cells:
        if cell == "KEY": columnIndexStore.keyIndex = currentIndex
        if cell == "正常模式": columnIndexStore.dayValueIndex = currentIndex
        if cell == "黑夜模式": columnIndexStore.nightValueIndex = currentIndex
        currentIndex = currentIndex + 1
    return columnIndexStore

# 在内存中创建一个空的文档
docDay = xml.dom.minidom.Document()
docNight = xml.dom.minidom.Document()
# 创建一个根节点resources对象
rootDay = docDay.createElement('resources')
rootNight = docNight.createElement('resources')
# 将根节点添加到文档对象中
docDay.appendChild(rootDay)
docNight.appendChild(rootNight)

try:
    wb = load_workbook(inputFilePath)
    # 获取workbook中所有的表格
    sheets = wb.sheetnames

    for sheetIndex in range(len(sheets)):
        sheet_name = sheets[sheetIndex]
        print("----------【" + sheet_name + "】页遍历开始---------------")
        rootDay.appendChild(docDay.createComment("【" + sheet_name + "】页遍历开始"))
        rootNight.appendChild(docNight.createComment("【" + sheet_name + "】页遍历开始"))

        sheet = wb[sheets[sheetIndex]]
        columnIndexStore = findSheetColumnIndex(sheet)
        if columnIndexStore.keyIndex == -1:
            print("----------【" + sheet_name + "】页遍历结束---------------")
            rootDay.appendChild(docDay.createComment("【" + sheet_name + "】页遍历结束"))
            rootNight.appendChild(docNight.createComment("【" + sheet_name + "】页遍历结束"))
            continue

        # 遍历所有行
        for rowIndex in range(2, sheet.max_row + 1):
            if sheet.cell(rowIndex, columnIndexStore.keyIndex).value is None: continue

            # todo openpyxl有bug，有时读取不到单元格的删除线
            if sheet.cell(rowIndex, columnIndexStore.keyIndex).font.strike or \
                    sheet.cell(rowIndex, columnIndexStore.dayValueIndex).font.strike or \
                    sheet.cell(rowIndex, columnIndexStore.nightValueIndex).font.strike:
                print(sheet.cell(rowIndex, columnIndexStore.keyIndex).value + "标上删除线，不做同步")
                continue

            dayValue = sheet.cell(rowIndex, columnIndexStore.dayValueIndex).value
            nightValue = sheet.cell(rowIndex, columnIndexStore.nightValueIndex).value
            if dayValue is not None:
                nodeDayColor = docDay.createElement('color')
                nodeDayColor.setAttribute('name', sheet.cell(rowIndex, columnIndexStore.keyIndex).value)
                nodeDayColor.appendChild(docDay.createTextNode(dayValue[0:9]))
                rootDay.appendChild(nodeDayColor)

            if nightValue is not None:
                nodeNightColor = docNight.createElement('color')
                nodeNightColor.setAttribute('name', sheet.cell(rowIndex, columnIndexStore.keyIndex).value)
                nodeNightColor.appendChild(docDay.createTextNode(nightValue[0:9]))
                rootNight.appendChild(nodeNightColor)

        print("----------【" + sheet_name + "】页遍历结束---------------")
        rootDay.appendChild(docDay.createComment("【" + sheet_name + "】页遍历结束"))
        rootNight.appendChild(docNight.createComment("【" + sheet_name + "】页遍历结束"))

        # 写入文件
        mkdir(outPutFileDir)
        fpDay = open(outPutFileDir + 'values.xml', 'w', encoding='utf-8')
        fpNight = open(outPutFileDir + 'values-night.xml', 'w', encoding='utf-8')
        docDay.writexml(fpDay, indent='', addindent='\t', newl='\n', encoding='utf-8')
        docNight.writexml(fpNight, indent='', addindent='\t', newl='\n', encoding='utf-8')
except FileNotFoundError as e:
    print("Excel file not found " + str(e))
    tkinter.messagebox.showinfo('提示', "找不到\"" + inputFilePath + "\"")
except Exception as e:
    print(str(e))